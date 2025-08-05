import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import numpy as np

def excel_toplam_ornegi():
    """
    Excel'de basit toplama işlemleri örneği
    """
    
    # Yeni bir Excel dosyası oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Toplama İşlemleri"
    
    # Başlıkları ekle
    basliklar = ["A", "B", "C", "Toplam"]
    for col, baslik in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=col, value=baslik)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Örnek veriler ekle
    veriler = [
        [10, 20, 30],
        [15, 25, 35],
        [5, 10, 15],
        [8, 12, 20],
        [12, 18, 25]
    ]
    
    # Verileri Excel'e yaz
    for row, veri in enumerate(veriler, 2):
        for col, deger in enumerate(veri, 1):
            ws.cell(row=row, column=col, value=deger)
        
        # Toplam formülü ekle
        toplam_formul = f"=SUM(A{row}:C{row})"
        ws.cell(row=row, column=4, value=toplam_formul)
    
    # Genel toplam satırı ekle
    son_satir = len(veriler) + 2
    ws.cell(row=son_satir, column=1, value="GENEL TOPLAM")
    ws.cell(row=son_satir, column=1).font = Font(bold=True)
    
    # Her sütun için genel toplam
    for col in range(2, 5):
        if col == 4:  # Toplam sütunu
            genel_toplam_formul = f"=SUM(D2:D{son_satir-1})"
        else:
            genel_toplam_formul = f"=SUM({chr(64+col)}2:{chr(64+col)}{son_satir-1})"
        
        ws.cell(row=son_satir, column=col, value=genel_toplam_formul)
        ws.cell(row=son_satir, column=col).font = Font(bold=True)
    
    # Sütun genişliklerini ayarla
    for col in range(1, 5):
        ws.column_dimensions[chr(64+col)].width = 15
    
    # Dosyayı kaydet
    dosya_adi = "toplama_islemi.xlsx"
    wb.save(dosya_adi)
    print(f"Excel dosyası '{dosya_adi}' oluşturuldu!")
    
    return dosya_adi

def pandas_ile_toplam():
    """
    Pandas kullanarak Excel'de toplama işlemi
    """
    
    # Örnek veri oluştur
    data = {
        'A': [10, 15, 5, 8, 12],
        'B': [20, 25, 10, 12, 18],
        'C': [30, 35, 15, 20, 25]
    }
    
    df = pd.DataFrame(data)
    
    # Toplam sütunu ekle
    df['Toplam'] = df.sum(axis=1)
    
    # Genel toplam satırı ekle
    genel_toplam = df.sum()
    genel_toplam_dict = genel_toplam.to_dict()
    genel_toplam_dict['A'] = 'GENEL TOPLAM'
    df = pd.concat([df, pd.DataFrame([genel_toplam_dict])], ignore_index=True)
    
    # Excel dosyasına kaydet
    dosya_adi = "pandas_toplam.xlsx"
    with pd.ExcelWriter(dosya_adi, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Toplam', index=False)
        
        # Çalışma sayfasını al
        worksheet = writer.sheets['Toplam']
        
        # Başlık formatını ayarla
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    print(f"Pandas ile Excel dosyası '{dosya_adi}' oluşturuldu!")
    return dosya_adi

def basit_hesap_makinesi():
    """
    Basit hesap makinesi fonksiyonu
    """
    print("=== Basit Hesap Makinesi ===")
    
    while True:
        print("\nİşlem seçin:")
        print("1. Toplama")
        print("2. Çıkarma")
        print("3. Çarpma")
        print("4. Bölme")
        print("5. Çıkış")
        
        secim = input("Seçiminiz (1-5): ")
        
        if secim == '5':
            print("Hesap makinesi kapatılıyor...")
            break
        
        if secim not in ['1', '2', '3', '4']:
            print("Geçersiz seçim! Lütfen 1-5 arası bir sayı girin.")
            continue
        
        try:
            sayi1 = float(input("Birinci sayıyı girin: "))
            sayi2 = float(input("İkinci sayıyı girin: "))
            
            if secim == '1':
                sonuc = sayi1 + sayi2
                islem = "Toplama"
            elif secim == '2':
                sonuc = sayi1 - sayi2
                islem = "Çıkarma"
            elif secim == '3':
                sonuc = sayi1 * sayi2
                islem = "Çarpma"
            elif secim == '4':
                if sayi2 == 0:
                    print("Hata: Sıfıra bölme hatası!")
                    continue
                sonuc = sayi1 / sayi2
                islem = "Bölme"
            
            print(f"{islem} sonucu: {sayi1} {'+' if secim == '1' else '-' if secim == '2' else '*' if secim == '3' else '/'} {sayi2} = {sonuc}")
            
        except ValueError:
            print("Hata: Lütfen geçerli bir sayı girin!")

if __name__ == "__main__":
    print("Excel Toplama İşlemleri Programı")
    print("=" * 40)
    
    # Excel dosyaları oluştur
    print("\n1. OpenPyXL ile Excel dosyası oluşturuluyor...")
    excel_toplam_ornegi()
    
    print("\n2. Pandas ile Excel dosyası oluşturuluyor...")
    pandas_ile_toplam()
    
    print("\n3. Basit hesap makinesi başlatılıyor...")
    basit_hesap_makinesi()